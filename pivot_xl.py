import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

if __name__ == '__main__':
    excel_file = pd.read_excel('Data/sales_supermarket.xlsx')
    excel_file[['Gender', 'Product line', 'Total']]

    report_table = excel_file.pivot_table(index='Gender',
                                          columns='Product line',
                                          values='Total',
                                          aggfunc='sum').round(0)

    report_table.to_excel('report_2021.xlsx',
                          sheet_name='Report',
                          startrow=4)

    wb = load_workbook('report_2021.xlsx')
    sheet = wb['Report']

    # cell references (original spreadsheet)
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    print(f'Min Columns: {min_column}')
    print(f'Max Columns: {max_column}')
    print(f'Min Rows: {min_row}')
    print(f'Max Rows: {max_row}')